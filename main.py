import re
from PIL import Image
from autocrop import Cropper
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject,
                            QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent, QThread, Signal, Slot)
from PySide2.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase,
                           QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient, QImage)
from PySide2.QtMultimedia import *
from PySide2.QtMultimediaWidgets import *
from PySide2.QtWidgets import *

import sys
import random
import string
import hashlib
import pathlib
import os
import mariadb
import plotly.express as px

from openpyxl import load_workbook

# GUI FILES
from ui_classes.ui_main_ import Ui_MainWindow
from ui_classes.signInFailedOneDialog import Ui_signInFailedOneDialog
from ui_classes.ui_incorrectDialog import Ui_incorrectDialog
import ui_classes.edit as edit
from ui_classes.edit import Ui_Dialog


# IMPORT FUNCTIONS
from ui_functions import *

BASE_DIR = pathlib.Path(os.path.dirname(os.path.abspath(__file__)))
FOLDER_NAME = '.TEMP'

SAVE_PATH = BASE_DIR / FOLDER_NAME
SAVE_PATH.mkdir(exist_ok=True, parents=True)

CONNECTION = mariadb.connect(
    host="localhost",
    user="root",
    passwd="",
    database="wassceverse",
    port=3306
)

CURSOR = CONNECTION.cursor()

class PasswordIncorrectDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_incorrectDialog()
        self.ui.setupUi(self)


class FailDialogOne(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_signInFailedOneDialog()
        self.ui.setupUi(self)


class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.availableCameras = QCameraInfo.availableCameras()
        self.viewFinder = self.ui.camera_input
        self.savePath = None
        self.school_id = 0
        self.bece_year = 2019
        self.students = []
        self.studentData = []
        self.registered_sudents_list = []
        self.currentStudentId = 0
        self.studentsNo = 0
        self.ui.stackedWidget.setCurrentIndex(0)
        self.ui.comboBox.addItems([camera.description()
                                  for camera in self.availableCameras])

        sql = "SELECT * FROM registered_schools"
        CURSOR.execute(sql)
        self.schools = []
        self.codes = []
        self.emails = []
        self.edit_add = ""
        for data in CURSOR:
            self.schools.append(data[1])
            self.codes.append(data[4])
            self.emails.append(data[6])

        completer = QCompleter(self.schools)
        self.ui.schoolNameSignIn.setCompleter(completer)

        def camera_screen():
            self.selectCamera(0)
            self.ui.stackedWidget.setCurrentIndex(3)

        def generate_key():
            index_numbers = []
            sql = "SELECT index_number FROM student_details WHERE school = ? AND bece_year = ?"
            CURSOR.execute(sql, (self.school_id, self.bece_year))
            for d in CURSOR:
                index_numbers.append(d[0])
            
            for index_number in index_numbers:
                sql = "UPDATE student_details SET student_key = ? WHERE index_number = ?"
                letters = string.ascii_lowercase
                code = ''.join(random.choice(letters) for i in range(6))
                CURSOR.execute(sql,(code, index_number))
                CONNECTION.commit()


        # CHECK INFO
        def signIn():
            schoolName = self.ui.schoolNameSignIn.text()
            email = self.ui.emailSignIn.text()
            password = self.ui.passwordSignIn.text()
            pwd = hashlib.md5(password.encode("utf-8")).hexdigest()

            sql = "SELECT * FROM registered_schools WHERE school_name = ? AND school_email = ?"
            CURSOR.execute(sql, (schoolName, email))
            data = []
            for d in CURSOR:
                data.append(d)

            if len(data) != 0:
                if pwd != data[0][7]:
                    dialog = PasswordIncorrectDialog(self)
                    dialog.exec()
                else:
                    self.school_id = data[0][0]
                    switch_screen(1)
                    year_group = int(self.ui.year_group.currentText())

                    sql = "SELECT COUNT(*) AS count_students FROM student_details WHERE bece_year = ? and school = ?"
                    CURSOR.execute(sql, (year_group, self.school_id))
                    for data in CURSOR:
                        num_of_students = data[0]

                    sql = "SELECT COUNT(*) AS count_students FROM registered_students INNER JOIN student_details ON registered_students.student = student_details.id WHERE bece_year = ? and school = ?"
                    CURSOR.execute(sql, (year_group, self.school_id))
                    for data in CURSOR:
                        num_registered = data[0]
                    
                    sql = "SELECT COUNT(*) AS count_students FROM registered_students INNER JOIN student_details ON registered_students.student = student_details.id WHERE bece_year = ? and school = ? and cleared = 1"
                    CURSOR.execute(sql, (year_group, self.school_id))
                    for data in CURSOR:
                        num_cleared = data[0]

                    sql = "SELECT * from registered_students"
                    CURSOR.execute(sql)
                    for data in CURSOR:
                        self.registered_sudents_list.append(data[1])

                    self.ui.label_40.setText(f"Number of Students: {num_of_students}")
                    self.ui.label_43.setText(f"Number of Students Registered: {num_registered}")
                    self.ui.label_44.setText(f"Number of Students Cleared: {num_cleared}")

                    

                    sql = "SELECT * FROM student_details WHERE school = ?"
                    CURSOR.execute(sql, (self.school_id,))
                    data = []
                    for d in CURSOR:
                        data.append(d)
                    self.studentsNo = len(data)
                    self.bece_year = self.ui.year_group.currentText()
                    sql = "SELECT bece_year FROM student_details where school = ?"
                    CURSOR.execute(sql, (self.school_id,))
                    years = []
                    for d in CURSOR:
                        years.append(d[0])
                    years.sort()

                    total_num = []
                    for year in years:
                        sql = "SELECT COUNT(*) FROM student_details WHERE student_details.bece_year = ? AND student_details.school = ?"
                        CURSOR.execute(sql, (year, self.school_id))
                        for d in CURSOR:
                            total_num.append(d[0])
                    
                    reg_list = []
                    for year in years:
                        sql = "SELECT COUNT(*) FROM registered_students JOIN student_details ON student_details.id = registered_students.student WHERE student_details.bece_year = ? AND student_details.school = ?"
                        CURSOR.execute(sql, (year, self.school_id))
                        for d in CURSOR:
                            reg_list.append(d[0])

                    clear_list = []
                    for year in years:
                        sql = "SELECT COUNT(*) FROM registered_students JOIN student_details ON student_details.id = registered_students.student WHERE student_details.bece_year = ? AND student_details.school = ? AND cleared = 1"
                        CURSOR.execute(sql, (year, self.school_id))
                        for d in CURSOR:
                            clear_list.append(d[0])

                    self.ui.widget_2.setHtml(f"""
                        <html>
                            <head>
                                <script src='https://cdn.plot.ly/plotly-2.16.1.min.js'></script>
                            </head>
                            <body>
                                <div id="graph" style="height: 400px;"></div>
                                <script>
                                    var trace1 = {{
                                        x: {years},
                                        y: {total_num},
                                        name: 'Total Number of Students',
                                        type: 'bar'
                                    }};

                                    var trace2 = {{
                                        x: {years},
                                        y: {reg_list},
                                        name: 'Registered Students',
                                        type: 'bar'
                                    }};

                                    var trace3 = {{
                                        x: {years},
                                        y: {clear_list},
                                        name: 'Students Cleared',
                                        type: 'bar'
                                    }};

                                    var data = [trace1, trace2, trace3];

                                    var layout = {{barmode: 'group'}};

                                    Plotly.newPlot('graph', data, layout);

                                </script>
                            </body>
                        </html>
                    """)

                    if self.studentsNo > 0:
                        self.getStudent(0)
                        self.ui.view_data.clicked.connect(lambda: switch_screen(4))
                    else:
                        self.ui.view_data.clicked.connect(lambda: switch_screen(7))
            else:
                dialog = FailDialogOne(self)
                dialog.exec()

        def signUp(db: str):
            school_name = self.ui.schoolNameSignUp.text()
            email = self.ui.emailSignUp.text()
            password = self.ui.passwordSignUp.text()
            location = self.ui.locationSignUp.text()
            country = self.ui.countrySignUp.text()

            if len(password) <= 6:
                self.ui.password_error.setHidden(False)
                self.ui.school_name_error.setHidden(True)
                self.ui.school_code_error.setHidden(True)
            elif len(email) > 0 and len(school_name) > 0 and len(location) > 0 and len(country) > 0:
                regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if (not re.search(regex, email)):
                    "TODO"
                elif email in self.emails:
                    self.ui.email_error.setHidden(False)
                    self.ui.school_name_error.setHidden(True)
                    self.ui.password_error.setHidden(True)
                elif school_name in self.schools:
                    self.ui.school_name_error.setHidden(False)
                    self.ui.email_error.setHidden(True)
                    self.ui.password_error.setHidden(True)
                else:
                    sql = "INSERT INTO registered_schools (school_name, location, country, school_email, password) VALUES (?, ?, ?, ?, MD5(?))"
                    CURSOR.execute(
                        sql, (school_name, location, country, email, password,))
                    CONNECTION.commit()
                    switch_screen(0)

        def nextStudent():
            id = self.currentStudentId
            if self.currentStudentId < self.studentsNo:
                id = self.currentStudentId + 1
            self.getStudent(id)

        def previousStudent():
            id = 0
            if self.currentStudentId > 1:
                id = self.currentStudentId - 1
            self.getStudent(id)

        def delete_student():
            sql = "DELETE FROM student_details WHERE id = ?"
            CURSOR.execute(sql, (self.studentData[0],))
            CONNECTION.commit()
            if self.currentStudentId + 1 == self.studentsNo:
                self.getStudent(self.currentStudentId-1)
            else:
                self.getStudent(self.currentStudentId)

        def update():
            surname = self.ui.plainTextEdit.toPlainText()
            first_name = self.ui.plainTextEdit_2.toPlainText()
            other_names = self.ui.plainTextEdit_3.toPlainText()
            dob = self.ui.dateEdit.date().toString("dd/MM/yyyy")
            course = self.ui.comboBox_2.currentText()
            class_ = self.ui.textEdit.toPlainText()
            index_number = self.ui.plainTextEdit_4.toPlainText()
            year_completed = str(self.ui.dateEdit_2.date().year())

            checkboxes = [
                self.ui.checkBox,
                self.ui.checkBox_2,
                self.ui.checkBox_3,
                self.ui.checkBox_4,
                self.ui.checkBox_5,
                self.ui.checkBox_6,
                self.ui.checkBox_7,
                self.ui.checkBox_8,
                self.ui.checkBox_9,
                self.ui.checkBox_10,
                self.ui.checkBox_11,
                self.ui.checkBox_12,
                self.ui.checkBox_13,
                self.ui.checkBox_14,
                self.ui.checkBox_15,
                self.ui.checkBox_16,
                self.ui.checkBox_17,
                self.ui.checkBox_18,
                self.ui.checkBox_19,
                self.ui.checkBox_20,
                self.ui.checkBox_21,
                self.ui.checkBox_22,
                self.ui.checkBox_23,
                self.ui.checkBox_24
            ]

            radios = [self.ui.radioButton, self.ui.radioButton_2]

            elective = []
            for checkbox in checkboxes:
                if checkbox.isChecked():
                    elective.append(checkbox.text())

            gender = ""
            for radio in radios:
                if radio.isChecked():
                    gender = radio.text()
            
            if gender.lower() == "male":
                gender = 0
            elif gender.lower() == "female":
                gender = 1

            print(gender)

            parent_contact = self.ui.plainTextEdit_5.toPlainText()
            if len(elective) == 4:
                electives = f'{elective[0]},{elective[1]},{elective[2]},{elective[3]}'
            else:
                electives = ""
            
            if self.edit_add == "add":
                if surname != "" or first_name != "" or other_names != "" or index_number != "" or class_ != "" or gender != "" or electives != "":
                    sql = "INSERT INTO student_details (school, surname, first_name, other_names, course, class, index_number, electives, gender, parent_contact, date_of_birth, bece_year) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                    CURSOR.execute(sql, (self.school_id, surname, first_name, other_names, course, class_, index_number, electives, gender, parent_contact, dob, year_completed))
                    CONNECTION.commit()

                    sql = "SELECT * FROM student_details WHERE surname = ? AND first_name = ? AND other_names = ? AND index_number = ?"
                    CURSOR.execute(sql, (surname, first_name, other_names, index_number,))
                    data = []
                    for d in CURSOR:
                        data.append(d)
                        student = data[0][0]
                    
                    sql = "INSERT INTO registered_students (student) VALUES (?)"
                    CURSOR.execute(sql, (student,))
                    CONNECTION.commit()
            
            elif self.edit_add == "edit":
                if surname != "" or first_name != "" or other_names != "" or index_number != "" or class_ != "" or gender != "" or electives != "":
                    sql = "UPDATE student_details SET school = ?, surname = ?, first_name = ?, other_names = ?, course = ?, class = ?, index_number = ?, electives = ?, gender = ?, parent_contact = ?, date_of_birth = ?, bece_year = ? WHERE id = ?"
                    CURSOR.execute(sql, (self.school_id, surname, first_name, other_names, course, class_, index_number, electives, gender, parent_contact, dob, year_completed, self.studentData[0]))
                    CONNECTION.commit()

            elif self.edit_add == "register":
                if surname != "" or first_name != "" or other_names != "" or index_number != "" or class_ != "" or gender != "" or electives != "":
                    sql = "UPDATE student_details SET school = ?, surname = ?, first_name = ?, other_names = ?, course = ?, class = ?, index_number = ?, electives = ?, gender = ?, parent_contact = ?, date_of_birth = ?, bece_year = ? WHERE id = ?"
                    CURSOR.execute(sql, (self.school_id, surname, first_name, other_names, course, class_, index_number, electives, gender, parent_contact, dob, year_completed, self.studentData[0]))
                    CONNECTION.commit()

                    sql = "SELECT COUNT(*) FROM registered_students WHERE student = ?"
                    CURSOR.execute(sql, (self.studentData[0],))
                    data = []
                    for d in CURSOR:
                        data.append(d)
                        student = data[0]
                    
                    if student == 0:
                        sql = "INSERT INTO registered_students (student) VALUES (?)"
                        CURSOR.execute(sql, (self.studentData[0],))
                        CONNECTION.commit()
            
            self.edit_add = ""
            switch_screen(4)

        def edit_student_function(action):
            switch_screen(6, action)
            school = self.school_id
            index = self.studentData[0]
            listIndex = self.currentStudentId
            sql = "SELECT * FROM student_details WHERE school = ? and id = ?"
            CURSOR.execute(sql, (school, index))

            result = []
            for data in CURSOR:
                result.append(data)

            checkboxes = [
                self.ui.checkBox,
                self.ui.checkBox_2,
                self.ui.checkBox_3,
                self.ui.checkBox_4,
                self.ui.checkBox_5,
                self.ui.checkBox_6,
                self.ui.checkBox_7,
                self.ui.checkBox_8,
                self.ui.checkBox_9,
                self.ui.checkBox_10,
                self.ui.checkBox_11,
                self.ui.checkBox_12,
                self.ui.checkBox_13,
                self.ui.checkBox_14,
                self.ui.checkBox_15,
                self.ui.checkBox_16,
                self.ui.checkBox_17,
                self.ui.checkBox_18,
                self.ui.checkBox_19,
                self.ui.checkBox_20,
                self.ui.checkBox_21,
                self.ui.checkBox_22,
                self.ui.checkBox_23,
                self.ui.checkBox_24
            ]

            electives = data[8].split(",")
            for checkbox in checkboxes:
                for elective in electives:
                    if checkbox.text() == elective:
                        checkbox.setChecked(True)

            radios = [self.ui.radioButton, self.ui.radioButton_2]
            if data[9] == 0:
                gender = "Male"
            elif data[9] == 1:
                gender = "Female"
            for radio in radios:
                if radio.text() == gender:
                    radio.setChecked(True)
            self.ui.plainTextEdit.setPlainText(data[2])
            self.ui.plainTextEdit_2.setPlainText(data[3])
            self.ui.plainTextEdit_3.setPlainText(data[4])
            self.ui.textEdit.setText(data[6])
            self.ui.plainTextEdit_4.setPlainText(data[7])
            self.ui.plainTextEdit_5.setPlainText(data[10])
            self.ui.comboBox_2.setCurrentText(data[5])
            date = data[11].split("/")
            self.ui.dateEdit.setDate(QDate(int(date[2]), int(date[1]), int(date[0])))
            
        def getImageFromFile():
            src = self.get_image_file()
            saveDir = SAVE_PATH / self.studentData[9]
            saveDir.mkdir(exist_ok=True, parents=True)
            saveFolder = saveDir / self.studentData[5]
            saveFolder.mkdir(exist_ok=True, parents=True)
            fileName = f'{self.studentData[1]} {self.studentData[2]} {self.studentData[3]}'.strip(
            ) + '.jpg'
            savePath = saveFolder / fileName
            pathlib.Path(src).rename(str(savePath))
            # Get a Numpy array of the cropped image
            img_url = str(savePath)

            cropper = Cropper(
                width=150,
                height=200,
                face_percent=40
            )
            cropped_array = cropper.crop(img_url)

            # Save the cropped image with PIL if a face was detected:
            try:
                if len(cropped_array) != 0:
                    cropped_image = Image.fromarray(cropped_array)
                    cropped_image.save(savePath)

            except TypeError:
                self.savePath = None

        self.ui.SignInSubmit_2.clicked.connect(lambda: signUp("server2.db"))
        self.ui.next_data_button.clicked.connect(lambda: nextStudent())
        self.ui.previous_data_button.clicked.connect(lambda: previousStudent())
        self.ui.search_student_button.clicked.connect(lambda: switch_screen(5))
        self.ui.close_search.clicked.connect(lambda: switch_screen(3))
        self.ui.take_photo.clicked.connect(lambda: camera_screen())
        self.ui.capture.clicked.connect(lambda: self.clickPhoto())
        self.ui.SignUpButton.clicked.connect(lambda: switch_screen(2))
        self.ui.SignUpButton_2.clicked.connect(lambda: switch_screen(0))
        self.ui.close_camera.clicked.connect(lambda: self.closeCamera())
        self.ui.SignInSubmit.clicked.connect(lambda: signIn())
        self.ui.edit_student_button.clicked.connect(lambda: edit_student_function("edit"))
        self.ui.add_student_button.clicked.connect(lambda: switch_screen(6, "add"))
        self.ui.pushButton.clicked.connect(lambda: switch_screen(6, "add"))
        self.ui.delete_student_button.clicked.connect(lambda: delete_student())
        self.ui.import_file.clicked.connect(lambda: getImageFromFile())
        self.ui.searchbar_main.textChanged.connect(lambda: getStudentList())
        self.ui.save_edit.clicked.connect(lambda: update())
        self.ui.pushButton_9.clicked.connect(lambda: generate_key())
        self.ui.cancel_edit.clicked.connect(lambda: switch_screen(4))
        self.ui.year_group.activated.connect(lambda: bece_year_update())
        self.ui.pushButton_5.clicked.connect(lambda: switch_screen(1))
        self.ui.pushButton_7.clicked.connect(lambda: self.logout())
        self.ui.pushButton_8.clicked.connect(lambda: edit_student_function("register"))
        self.ui.pushButton_6.clicked.connect(lambda: self.add_students())

        def bece_year_update():
            self.bece_year = self.ui.year_group.currentText()

        def getStudentList():
            sql = "SELECT * FROM student_details WHERE school = ? and bece_year = ?"
            CURSOR.execute(sql, (self.school_id, self.bece_year))
            data = []
            for d in CURSOR:
                data.append(d)
            search = self.ui.searchbar_main.text()
            names = []
            for d in data:
                name = f"{d[2]} {d[3]} {d[4]}".strip()
                names.append(name)

            # TODO: The search results
            self.ui.listWidget.clear()
            results = {}
            for n in names:
                if search in n:
                    results[n] = [n, names.index(n)]
                    self.ui.listWidget.addItem(n)

            def searchRes(item):
                self.getStudent(results[item.text()][1])
                self.ui.stackedWidget.setCurrentIndex(4)

            self.ui.listWidget.itemClicked.connect(searchRes)

        def switch_screen(screen: int, action: str = ""):
            self.edit_add = action
            self.ui.stackedWidget.setCurrentIndex(screen)

        # ==> SET UI DEFINITIONS
        UIFunctions.uiDefinitions(self)

        # SHOW ==> MAIN WINDOW
        ########################################################################
        self.show()

    # APP EVENTS
    ########################################################################

    def logout(self):
        self.ui.stackedWidget.setCurrentIndex(0)
        self.savePath = None
        self.school_id = 0
        self.bece_year = 2019
        self.students = []
        self.studentData = []
        self.currentStudentId = 0
        self.studentsNo = 0
        self.schools = []
        self.codes = []
        self.emails = []
        self.edit_add = ""
        self.ui.passwordSignIn.setText("")

    def get_image_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Select Image", "", "Image Files (*.jpeg *.jpg *.png)", options=options)
        return file_name

    def add_students(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Excel Spreadsheet (*.xlsx)", options=options)
        if fileName:
            workbook = load_workbook(filename=fileName)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=3, values_only=True):
                surname = row[0]
                first_name = row[1]
                other_names = row[2]
                course = row[4]
                class_ = row[3]
                index_number = row[5]
                electives = f"{row[6],row[7],row[8],row[9]}"
                gender = 0
                if row[10].lower == "male":
                    gender = 0
                elif row[10].lower == "female":
                    gender = 1

                parent_contact = row[11]
                dob = row[12]

                sql = "INSERT INTO student_details (school, surname, first_name, other_names, course, class, index_number, electives, gender, parent_contact, date_of_birth, bece_year) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                CURSOR.execute(sql, (self.school_id, surname, first_name, other_names, course, class_, index_number, electives, gender, parent_contact, dob, self.bece_year))
                CONNECTION.commit()


    def getStudent(self, id=0):
        sql = """SELECT
            student_details.id,
            student_details.surname,
            student_details.first_name,
            student_details.other_names,
            student_details.course,
            student_details.class,
            student_details.index_number,
            student_details.electives,
            student_details.gender,
            student_details.parent_contact,
            student_details.date_of_birth,
            registered_schools.school_name 
            FROM student_details JOIN registered_schools
            ON student_details.school = registered_schools.id
            WHERE student_details.school = ? AND student_details.bece_year = ?"""
        CURSOR.execute(sql, (self.school_id, self.bece_year))
        data = []
        for d in CURSOR:
            data.append(d)
        self.studentsNo = len(data)
        self.currentStudentId = id
        self.studentData = data[id]

        def previous_block():
            if self.currentStudentId == 0:
                self.ui.previous_data_button.setStyleSheet(
                    "background-color: rgb(200, 200, 200);"
                )
                self.ui.previous_data_button.setDisabled(True)
            else:
                self.ui.previous_data_button.setStyleSheet(
                    """QWidget {
                        background-color: rgb(112, 112, 112);
                        padding: 0px;
                        }

                        QWidget:hover {
                            background-color: rgb(168, 168, 168);
                        }
                    """
                )
                self.ui.previous_data_button.setDisabled(False)

        def next_block():
            if self.currentStudentId < self.studentsNo - 1:
                self.ui.next_data_button.setStyleSheet(
                    """QWidget {
                        background-color: rgb(112, 112, 112);
                        padding: 0px;
                        }

                        QWidget:hover {
                            background-color: rgb(168, 168, 168);
                        }
                    """
                )

                self.ui.next_data_button.setDisabled(False)

            elif self.currentStudentId == self.studentsNo - 1:
                self.ui.next_data_button.setStyleSheet(
                    "background-color: rgb(70, 70, 70);"
                )
                self.ui.next_data_button.setDisabled(True)

        next_block()
        previous_block()

        if len(data) != 0:
            name = f"{data[id][1]} {data[id][2]} {data[id][3]}"
            self.ui.student_name.setText(name.strip())
            self.ui.student_school.setText(data[id][10])
            self.ui.student_class.setText(data[id][5])
            self.ui.student_course.setText(data[id][4])
            if data[id][0] in self.registered_sudents_list:
                self.ui.label_48.setText("Registered: Yes")
                self.ui.pushButton_8.setHidden(True)
            else:
                self.ui.label_48.setText("Registered: No")
                self.ui.pushButton_8.setHidden(False)

            if data[id][8] == 0:
                self.ui.student_gender.setText("Male")
            elif data[id][8] == 1:
                self.ui.student_gender.setText("Female")
            date = data[id][10]
            date = date.split("/")
            day = int(date[0])
            suffix = ""
            if 11 <= (day % 100) <= 13:
                suffix = 'th'
            else:
                suffix = ['th', 'st', 'nd', 'rd', 'th'][min(day % 10, 4)]
            day = f"{str(day)}{suffix}"

            month = ["January", "February", "March", "April", "May", "June",
                     "July", "August", "September", "October", "November", "December"][int(date[1])-1]

            self.currentStudentId = id

            date = f"{day} {month} {date[2]}"
            self.ui.date_of_birth_label.setText(date)
            self.ui.parent_contact.setText(data[id][9])
            self.ui.index_number_bece.setText(data[id][6])
            electives_array = data[id][7].split(",")
            self.ui.elective_1.setText(electives_array[0])
            self.ui.elective_2.setText(electives_array[1])
            self.ui.elective_3.setText(electives_array[2])
            self.ui.elective_4.setText(electives_array[3])

            saveDir = SAVE_PATH / self.studentData[10]
            saveFolder = saveDir / self.studentData[5]
            fileName = f'{self.studentData[1]} {self.studentData[2]} {self.studentData[3]}'.strip(
            ) + '.jpg'
            savePath = saveFolder / fileName
            if savePath.exists():
                pixmap = QtGui.QPixmap(str(savePath))
            else:
                pixmap = QtGui.QPixmap(
                    'Include/img/user-sign-icon-person-symbol-human-avatar-vector-12693195.jpg')
            self.ui.label_12.setPixmap(pixmap)

            # FIXME: Getting image from database
            """ if type(data[id][14]) == bytes:
                if len(data[id][14]) > 0:
                    ba = QtCore.QByteArray(data[id][14])
                    pixmap = QtGui.QPixmap()
                    ok = pixmap.loadFromData(ba, "JPG")
                    assert ok
                else:
                    pixmap = QtGui.QPixmap('Include/img/user-sign-icon-person-symbol-human-avatar-vector-12693195.jpg')
                self.ui.label_12.setPixmap(pixmap) """

    def selectCamera(self, i):
        self.camera = QCamera(self.availableCameras[i])
        self.camera.setViewfinder(self.viewFinder)
        self.camera.setCaptureMode(QCamera.CaptureStillImage)
        self.camera.error.connect(lambda:
                                  self.alert(self.camera.errorString()))
        self.camera.start()
        self.capture = QCameraImageCapture(self.camera)
        self.capture.error.connect(lambda d, i:
                                   print(
                                       f'Image Captured {str(self.saveSeq)}'))
        self.currentCameraName = self.availableCameras[i].description()
        self.saveSeq = 0

    def clickPhoto(self):
        fileName = f'{self.studentData[1]} {self.studentData[2]} {self.studentData[3]}'.strip(
        ) + '.jpg'
        saveDir = SAVE_PATH / self.studentData[9]
        saveDir.mkdir(exist_ok=True, parents=True)
        saveFolder = saveDir / self.studentData[5]
        saveFolder.mkdir(exist_ok=True, parents=True)
        savePath = saveFolder / fileName
        self.savePath = savePath
        self.saveImage(str(savePath))
        print('Image saved on ', str(savePath))
        self.saveSeq += 1

    def closeCamera(self):
        if self.savePath == None:
            self.ui.stackedWidget.setCurrentIndex(3)
        else:
            cropper = Cropper(
                width=150,
                height=200,
                face_percent=40
            )

            # Get a Numpy array of the cropped image
            img_url = str(self.savePath)
            cropped_array = cropper.crop(img_url)

            # Save the cropped image with PIL if a face was detected:
            try:
                if len(cropped_array) != 0:
                    cropped_image = Image.fromarray(cropped_array)
                    cropped_image.save(self.savePath)
                with open(self.savePath, 'rb') as input_file:
                    ablob = input_file.read()
                    base = os.path.basename(self.savePath)
                    afile, ext = os.path.splitext(base)
                    sql = "UPDATE student_details SET image = ? WHERE index_number = ?"
                    CURSOR.execute(
                        sql, (mariadb.Binary(ablob), self.studentData[6]))
                    CONNECTION.commit()
                self.savePath = None
                self.getStudent(self.currentStudentId)
                self.ui.stackedWidget.setCurrentIndex(3)
            except TypeError:
                self.savePath = None
                self.ui.stackedWidget.setCurrentIndex(3)

    def saveImage(self, savePath: str):
        self.capture.capture(savePath)

    def alert(self, msg):
        error = QErrorMessage(self)
        error.showMessage(msg)

    def mousePressEvent(self, event):
        self.dragPos = event.globalPos()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec_())
